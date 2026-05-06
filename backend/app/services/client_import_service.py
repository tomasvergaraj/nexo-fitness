"""Bulk client import from XLSX/CSV files."""

from __future__ import annotations

import csv
import io
import json
import re
import secrets
import string
import uuid
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.models.user import User, UserRole

MAX_ROWS = 500
MAX_FILE_SIZE_BYTES = 2 * 1024 * 1024  # 2 MB
PREVIEW_TTL_SECONDS = 30 * 60  # 30 minutes
REDIS_KEY_PREFIX = "client_import:"

COLUMN_MAP: dict[str, str] = {
    "nombre": "first_name",
    "apellido": "last_name",
    "email": "email",
    "correo": "email",
    "telefono": "phone",
    "teléfono": "phone",
    "fecha de nacimiento": "date_of_birth",
    "fecha nacimiento": "date_of_birth",
    "género": "gender",
    "genero": "gender",
    "contacto emergencia": "emergency_contact",
    "telefono emergencia": "emergency_phone",
    "teléfono emergencia": "emergency_phone",
    "notas medicas": "medical_notes",
    "notas médicas": "medical_notes",
    "tags": "tags",
    "etiquetas": "tags",
}

REQUIRED_FIELDS = ("first_name", "last_name", "email")
GENDER_OPTIONS = {"masculino", "femenino", "otro"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@dataclass
class RowError:
    row: int
    column: str
    message: str


@dataclass
class ValidatedRow:
    row: int
    first_name: str
    last_name: str
    email: str
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None  # ISO YYYY-MM-DD; serialized for Redis
    gender: Optional[str] = None
    emergency_contact: Optional[str] = None
    emergency_phone: Optional[str] = None
    medical_notes: Optional[str] = None
    tags: list[str] = field(default_factory=list)


@dataclass
class ImportPreview:
    import_token: str
    total_rows: int
    valid_count: int
    error_count: int
    valid_preview: list[dict[str, Any]]
    errors: list[dict[str, Any]]
    quota_remaining: int
    quota_max: int
    quota_blocked: bool


def generate_password() -> str:
    """Generate a 12-char password matching the validator (>=1 upper, >=1 digit)."""
    alphabet = string.ascii_letters + string.digits
    while True:
        candidate = "".join(secrets.choice(alphabet) for _ in range(12))
        if any(c.isupper() for c in candidate) and any(c.isdigit() for c in candidate):
            return candidate


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = [
        "Nombre",
        "Apellido",
        "Email",
        "Teléfono",
        "Fecha de nacimiento",
        "Género",
        "Contacto emergencia",
        "Teléfono emergencia",
        "Notas médicas",
        "Tags",
    ]
    example = [
        "Camila",
        "Pérez",
        "camila.perez@ejemplo.cl",
        "+56 9 1234 5678",
        "1995-08-22",
        "femenino",
        "Juan Pérez",
        "+56 9 8765 4321",
        "Asma leve",
        "vip,crossfit",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = max(18, len(header) + 2)

    for col, value in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=value)

    ws.cell(row=3, column=1, value="").comment = None  # placeholder

    gender_validation = DataValidation(
        type="list",
        formula1='"masculino,femenino,otro"',
        allow_blank=True,
    )
    gender_validation.error = "Usa masculino, femenino u otro"
    gender_validation.errorTitle = "Género inválido"
    ws.add_data_validation(gender_validation)
    gender_validation.add(f"F2:F{MAX_ROWS + 1}")

    notes_ws = wb.create_sheet("Instrucciones")
    instructions = [
        "Cómo usar esta plantilla",
        "",
        "1. Completa una fila por cliente. La fila 2 es un ejemplo: bórrala antes de importar.",
        "2. Las columnas Nombre, Apellido y Email son obligatorias.",
        "3. La fecha de nacimiento puede ir en formato AAAA-MM-DD o DD/MM/AAAA.",
        "4. Tags se separan con coma (ej: vip,crossfit).",
        "5. Cada cliente recibe una contraseña aleatoria — comparte con ellos el flujo de recuperar contraseña.",
        f"6. Máximo {MAX_ROWS} filas por archivo, peso máximo 2 MB.",
        "",
        "Después de subir verás un resumen y podrás corregir errores antes de confirmar.",
    ]
    for idx, line in enumerate(instructions, start=1):
        cell = notes_ws.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
    notes_ws.column_dimensions["A"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_template_csv() -> bytes:
    headers = [
        "Nombre",
        "Apellido",
        "Email",
        "Teléfono",
        "Fecha de nacimiento",
        "Género",
        "Contacto emergencia",
        "Teléfono emergencia",
        "Notas médicas",
        "Tags",
    ]
    example = [
        "Camila",
        "Pérez",
        "camila.perez@ejemplo.cl",
        "+56 9 1234 5678",
        "1995-08-22",
        "femenino",
        "Juan Pérez",
        "+56 9 8765 4321",
        "Asma leve",
        "vip,crossfit",
    ]
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerow(example)
    return ("﻿" + buffer.getvalue()).encode("utf-8")


EXPORT_HEADERS = [
    "Nombre",
    "Apellido",
    "Email",
    "Teléfono",
    "Fecha de nacimiento",
    "Género",
    "Contacto emergencia",
    "Teléfono emergencia",
    "Notas médicas",
    "Tags",
    "Plan actual",
    "Estado membresía",
    "Vence el",
    "Estado cliente",
    "Riesgo de baja",
    "Último check-in",
    "Fecha de registro",
]


def _format_date_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _build_export_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for r in records:
        rows.append([
            r.get("first_name") or "",
            r.get("last_name") or "",
            r.get("email") or "",
            r.get("phone") or "",
            _format_date_value(r.get("date_of_birth")),
            r.get("gender") or "",
            r.get("emergency_contact") or "",
            r.get("emergency_phone") or "",
            r.get("medical_notes") or "",
            ", ".join(r["tags"]) if r.get("tags") else "",
            r.get("plan_name") or "",
            r.get("membership_status") or "",
            _format_date_value(r.get("membership_expires_at")),
            "Activo" if r.get("is_active") else "Inactivo",
            r.get("churn_risk_label") or "",
            _format_date_value(r.get("last_checkin_at")),
            _format_date_value(r.get("created_at")),
        ])
    return rows


def build_export_xlsx(records: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_values in enumerate(_build_export_rows(records), start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns based on content (with sane caps)
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        column_cells = ws[letter]
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        ws.column_dimensions[letter].width = min(max(len(header) + 2, max_len + 2), 40)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_export_csv(records: list[dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS)
    for row in _build_export_rows(records):
        writer.writerow(row)
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def build_errors_xlsx(errors: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Errores"
    headers = ["Fila", "Columna", "Motivo"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = 30
    for idx, err in enumerate(errors, start=2):
        ws.cell(row=idx, column=1, value=err.get("row"))
        ws.cell(row=idx, column=2, value=err.get("column"))
        ws.cell(row=idx, column=3, value=err.get("message"))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _coerce_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, (int, float)):
        return str(value).strip()
    return str(value).strip() or None


def _parse_date(value: Any) -> tuple[Optional[date], Optional[str]]:
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = str(value).strip()
    if not text:
        return None, None
    formats = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, "Fecha inválida. Usa AAAA-MM-DD o DD/MM/AAAA"


def _parse_tags(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    raw = str(value)
    parts = [tag.strip() for tag in raw.split(",")]
    return [tag for tag in parts if tag]


def parse_workbook(file_bytes: bytes, filename: str) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, data_rows) parsed from XLSX or CSV."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], []
        headers = [_normalize_header(h) for h in rows[0]]
        return headers, rows[1:]

    # default to XLSX
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}") from exc
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []
    headers = [_normalize_header(h) for h in header_row]
    data: list[list[Any]] = []
    for row in iterator:
        if row is None:
            continue
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        data.append(list(row))
    return headers, data


def _resolve_field_indices(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        canonical = COLUMN_MAP.get(header)
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


async def _existing_emails(db: AsyncSession, emails: set[str]) -> set[str]:
    if not emails:
        return set()
    result = await db.execute(select(User.email).where(User.email.in_(emails)))
    return {row[0].lower() for row in result.all() if row[0]}


def validate_rows(
    headers: list[str],
    data_rows: list[list[Any]],
    *,
    existing_emails: set[str],
) -> tuple[list[ValidatedRow], list[RowError]]:
    field_index = _resolve_field_indices(headers)
    errors: list[RowError] = []
    valid: list[ValidatedRow] = []

    missing_required = [f for f in REQUIRED_FIELDS if f not in field_index]
    if missing_required:
        labels = ", ".join(missing_required)
        errors.append(RowError(row=1, column="encabezados", message=f"Faltan columnas obligatorias: {labels}"))
        return valid, errors

    seen_emails: dict[str, int] = {}

    for offset, row_values in enumerate(data_rows):
        excel_row = offset + 2  # header is row 1

        def cell(field_name: str) -> Any:
            idx = field_index.get(field_name)
            if idx is None or idx >= len(row_values):
                return None
            return row_values[idx]

        first_name = _coerce_str(cell("first_name"))
        last_name = _coerce_str(cell("last_name"))
        email_raw = _coerce_str(cell("email"))

        row_errors: list[RowError] = []

        if not first_name:
            row_errors.append(RowError(row=excel_row, column="Nombre", message="Nombre es obligatorio"))
        if not last_name:
            row_errors.append(RowError(row=excel_row, column="Apellido", message="Apellido es obligatorio"))

        email_normalized: Optional[str] = None
        if not email_raw:
            row_errors.append(RowError(row=excel_row, column="Email", message="Email es obligatorio"))
        else:
            email_normalized = email_raw.lower()
            if not EMAIL_RE.match(email_normalized):
                row_errors.append(RowError(row=excel_row, column="Email", message="Formato de email inválido"))
            elif email_normalized in existing_emails:
                row_errors.append(
                    RowError(
                        row=excel_row,
                        column="Email",
                        message="Email ya existe en tu base de clientes",
                    )
                )
            elif email_normalized in seen_emails:
                row_errors.append(
                    RowError(
                        row=excel_row,
                        column="Email",
                        message=f"Email duplicado en el archivo (también en fila {seen_emails[email_normalized]})",
                    )
                )

        dob_value, dob_error = _parse_date(cell("date_of_birth"))
        if dob_error:
            row_errors.append(RowError(row=excel_row, column="Fecha de nacimiento", message=dob_error))

        gender_raw = _coerce_str(cell("gender"))
        gender_value: Optional[str] = None
        if gender_raw:
            gender_value = gender_raw.lower()
            if gender_value not in GENDER_OPTIONS:
                row_errors.append(
                    RowError(
                        row=excel_row,
                        column="Género",
                        message="Usa masculino, femenino u otro",
                    )
                )
                gender_value = None

        if row_errors:
            errors.extend(row_errors)
            continue

        assert first_name and last_name and email_normalized

        seen_emails[email_normalized] = excel_row

        valid.append(
            ValidatedRow(
                row=excel_row,
                first_name=first_name,
                last_name=last_name,
                email=email_normalized,
                phone=_coerce_str(cell("phone")),
                date_of_birth=dob_value.isoformat() if dob_value else None,
                gender=gender_value,
                emergency_contact=_coerce_str(cell("emergency_contact")),
                emergency_phone=_coerce_str(cell("emergency_phone")),
                medical_notes=_coerce_str(cell("medical_notes")),
                tags=_parse_tags(cell("tags")),
            )
        )

    return valid, errors


async def _get_redis():
    import redis.asyncio as aioredis
    return aioredis.from_url(get_settings().REDIS_URL, decode_responses=True)


async def store_preview(tenant_id: uuid.UUID, rows: list[ValidatedRow]) -> str:
    token = uuid.uuid4().hex
    payload = {
        "tenant_id": str(tenant_id),
        "rows": [asdict(row) for row in rows],
    }
    redis = await _get_redis()
    try:
        await redis.set(
            f"{REDIS_KEY_PREFIX}{token}",
            json.dumps(payload),
            ex=PREVIEW_TTL_SECONDS,
        )
    finally:
        await redis.aclose()
    return token


async def load_preview(tenant_id: uuid.UUID, token: str) -> Optional[list[ValidatedRow]]:
    redis = await _get_redis()
    try:
        raw = await redis.get(f"{REDIS_KEY_PREFIX}{token}")
    finally:
        await redis.aclose()
    if not raw:
        return None
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return None
    if payload.get("tenant_id") != str(tenant_id):
        return None
    return [ValidatedRow(**item) for item in payload.get("rows", [])]


async def discard_preview(token: str) -> None:
    redis = await _get_redis()
    try:
        await redis.delete(f"{REDIS_KEY_PREFIX}{token}")
    finally:
        await redis.aclose()


def row_to_user_kwargs(row: ValidatedRow, tenant_id: uuid.UUID) -> dict[str, Any]:
    dob: Optional[datetime] = None
    if row.date_of_birth:
        try:
            dob = datetime.fromisoformat(row.date_of_birth)
        except ValueError:
            dob = None
    return {
        "tenant_id": tenant_id,
        "email": row.email,
        "first_name": row.first_name,
        "last_name": row.last_name,
        "phone": row.phone,
        "role": UserRole.CLIENT,
        "date_of_birth": dob,
        "gender": row.gender,
        "emergency_contact": row.emergency_contact,
        "emergency_phone": row.emergency_phone,
        "medical_notes": row.medical_notes,
        "tags": json.dumps(row.tags) if row.tags else None,
    }
